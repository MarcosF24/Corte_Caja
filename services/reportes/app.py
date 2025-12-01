import os
from io import BytesIO
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo   # <-- NUEVO

from flask import Flask, request, jsonify
import pymysql
from pymysql.cursors import DictCursor
import requests

# .env solo en local
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

import boto3
from openpyxl import Workbook
from fpdf import FPDF

app = Flask(__name__)

mx_tz = ZoneInfo("America/Mexico_City")   # <-- NUEVO

DB_CONFIG = {
    "host": os.getenv("DB_HOST", "localhost"),
    "user": os.getenv("DB_USER", "root"),
    "password": os.getenv("DB_PASS", ""),
    "database": os.getenv("DB_NAME", "corte_caja"),
    "cursorclass": DictCursor,
}

S3_REPORTES_BUCKET = os.getenv("S3_REPORTES_BUCKET")
NOTIFICACIONES_URL = os.getenv("NOTIFICACIONES_URL")


# ==========================
# Helpers de infraestructura
# ==========================

def get_connection():
    return pymysql.connect(**DB_CONFIG)


def subir_a_s3(nombre_objeto, contenido_bytes, content_type):
    if not S3_REPORTES_BUCKET:
        raise RuntimeError("S3_REPORTES_BUCKET no está configurado")

    s3 = boto3.client("s3")
    s3.put_object(
        Bucket=S3_REPORTES_BUCKET,
        Key=nombre_objeto,
        Body=contenido_bytes,
        ContentType=content_type,
    )

    return f"https://{S3_REPORTES_BUCKET}.s3.amazonaws.com/{nombre_objeto}"


# ==========================
# Lógica de negocio
# ==========================

def obtener_corte_final(corte_final_id):
    sql = """
        SELECT id, usuario_id, fecha_inicio, fecha_fin, turno, tipo_corte
        FROM cortes
        WHERE id = %s
    """
    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute(sql, (corte_final_id,))
            corte = cursor.fetchone()
    finally:
        conn.close()

    if not corte:
        return None

    if corte.get("tipo_corte") and corte["tipo_corte"] != "FINAL":
        return None

    return corte


def obtener_ultimo_corte_final_anterior(fecha_final):
    sql = """
        SELECT id, fecha_inicio
        FROM cortes
        WHERE tipo_corte = 'FINAL'
          AND fecha_inicio < %s
        ORDER BY fecha_inicio DESC
        LIMIT 1
    """
    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute(sql, (fecha_final,))
            return cursor.fetchone()
    finally:
        conn.close()


def obtener_cortes_turno_en_rango(fecha_desde, fecha_hasta):
    sql = """
        SELECT id, usuario_id, fecha_inicio, fecha_fin, turno
        FROM cortes
        WHERE tipo_corte = 'TURNO'
          AND fecha_inicio > %s
          AND fecha_inicio <= %s
        ORDER BY fecha_inicio ASC
    """
    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute(sql, (fecha_desde, fecha_hasta))
            return cursor.fetchall()
    finally:
        conn.close()


def calcular_totales_para_cortes(cortes_turno):
    if not cortes_turno:
        return {
            "ventas_efectivo": 0.0,
            "ventas_tarjeta": 0.0,
            "gastos": 0.0,
            "neto": 0.0,
        }

    corte_ids = [c["id"] for c in cortes_turno]
    placeholders = ",".join(["%s"] * len(corte_ids))

    sql = f"""
        SELECT
            SUM(CASE WHEN m.tipo = 'INGRESO' AND m.descripcion = 'VENTAS_EFECTIVO' THEN m.monto ELSE 0 END) AS ventas_efectivo,
            SUM(CASE WHEN m.tipo = 'INGRESO' AND m.descripcion = 'VENTAS_TARJETA' THEN m.monto ELSE 0 END) AS ventas_tarjeta,
            SUM(CASE WHEN m.tipo = 'EGRESO' THEN m.monto ELSE 0 END) AS gastos
        FROM movimientos m
        WHERE m.corte_id IN ({placeholders})
    """

    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute(sql, corte_ids)
            row = cursor.fetchone() or {}
    finally:
        conn.close()

    ventas_efectivo = float(row.get("ventas_efectivo") or 0)
    ventas_tarjeta = float(row.get("ventas_tarjeta") or 0)
    gastos = float(row.get("gastos") or 0)
    neto = ventas_efectivo + ventas_tarjeta - gastos

    return {
        "ventas_efectivo": ventas_efectivo,
        "ventas_tarjeta": ventas_tarjeta,
        "gastos": gastos,
        "neto": neto,
    }


def generar_pdf(data):
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Título
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "Reporte Final de Corte de Caja", ln=True)
    pdf.ln(4)

    # Datos generales
    pdf.set_font("Helvetica", "", 11)
    pdf.cell(0, 8, f"Fecha de reporte: {data['fecha_reporte']}", ln=True)
    pdf.cell(0, 8, f"Corte final ID: {data['corte_final_id']}", ln=True)
    pdf.cell(0, 8, f"Rango: {data['rango_desde']} a {data['rango_hasta']}", ln=True)
    pdf.ln(4)

    # Totales
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "Totales:", ln=True)
    pdf.set_font("Helvetica", "", 11)
    pdf.cell(0, 8, f"Ventas en efectivo: ${data['totales']['ventas_efectivo']:.2f}", ln=True)
    pdf.cell(0, 8, f"Ventas con tarjeta: ${data['totales']['ventas_tarjeta']:.2f}", ln=True)
    pdf.cell(0, 8, f"Gastos: ${data['totales']['gastos']:.2f}", ln=True)
    pdf.cell(0, 8, f"Neto: ${data['totales']['neto']:.2f}", ln=True)
    pdf.ln(4)

    # Resumen de cortes por turno
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, f"Cortes por turno incluidos: {len(data['cortes_turno'])}", ln=True)

    out = pdf.output(dest="S")
    if isinstance(out, (bytes, bytearray)):
        pdf_bytes = bytes(out)
    else:
        pdf_bytes = str(out).encode("latin-1")

    return pdf_bytes


def generar_excel(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Final"

    ws["A1"] = "Reporte Final de Corte de Caja"
    ws["A2"] = f"Fecha de reporte: {data['fecha_reporte']}"
    ws["A3"] = f"Corte final ID: {data['corte_final_id']}"
    ws["A4"] = f"Rango: {data['rango_desde']} a {data['rango_hasta']}"

    ws["A6"] = "Ventas efectivo"
    ws["B6"] = data["totales"]["ventas_efectivo"]
    ws["A7"] = "Ventas tarjeta"
    ws["B7"] = data["totales"]["ventas_tarjeta"]
    ws["A8"] = "Gastos"
    ws["B8"] = data["totales"]["gastos"]
    ws["A9"] = "Neto"
    ws["B9"] = data["totales"]["neto"]

    ws["A11"] = "Cortes por turno incluidos"
    ws["A12"] = "ID corte"
    ws["B12"] = "Usuario ID"
    ws["C12"] = "Fecha inicio"
    ws["D12"] = "Fecha fin"
    ws["E12"] = "Turno"

    row = 13
    for c_info in data["cortes_turno"]:
        ws[f"A{row}"] = c_info["id"]
        ws[f"B{row}"] = c_info["usuario_id"]
        ws[f"C{row}"] = c_info["fecha_inicio"].strftime("%Y-%m-%d %H:%M:%S") if c_info["fecha_inicio"] else ""
        ws[f"D{row}"] = c_info["fecha_fin"].strftime("%Y-%m-%d %H:%M:%S") if c_info["fecha_fin"] else ""
        ws[f"E{row}"] = c_info.get("turno") or ""
        row += 1

    buffer = BytesIO()
    wb.save(buffer)
    excel_bytes = buffer.getvalue()
    buffer.close()
    return excel_bytes


def guardar_reporte_bd(corte_final_id, pdf_url, excel_url):
    sql = """
        INSERT INTO reportes (corte_id, archivo_pdf_url, archivo_excel_url)
        VALUES (%s, %s, %s)
    """
    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute(sql, (corte_final_id, pdf_url, excel_url))
            conn.commit()
            return cursor.lastrowid
    finally:
        conn.close()


# ==========================
# Rutas / Endpoints
# ==========================

@app.route("/health", methods=["GET"])
def health():
    return jsonify({"status": "ok", "service": "reportes"}), 200


@app.route("/reportes/generar-desde-corte-final", methods=["POST"])
def generar_desde_corte_final():
    data = request.get_json(force=True) or {}
    corte_final_id = data.get("corte_final_id")

    if not corte_final_id:
        return jsonify({"error": "corte_final_id es requerido"}), 400

    corte_final = obtener_corte_final(corte_final_id)
    if not corte_final:
        return jsonify({"error": "Corte final no encontrado o no es tipo FINAL"}), 404

    fecha_final = corte_final["fecha_inicio"]

    ultimo_final = obtener_ultimo_corte_final_anterior(fecha_final)
    if ultimo_final:
        fecha_desde = ultimo_final["fecha_inicio"]
    else:
        fecha_desde = datetime(
            fecha_final.year, fecha_final.month, fecha_final.day, 0, 0, 0
        ).replace(tzinfo=mx_tz)

    cortes_turno = obtener_cortes_turno_en_rango(fecha_desde, fecha_final)

    cortes_para_totales = list(cortes_turno) + [corte_final]
    totales = calcular_totales_para_cortes(cortes_para_totales)

    payload_reporte = {
        "fecha_reporte": datetime.now(mx_tz).strftime("%Y-%m-%d %H:%M:%S"),
        "corte_final_id": corte_final_id,
        "rango_desde": fecha_desde.strftime("%Y-%m-%d %H:%M:%S"),
        "rango_hasta": fecha_final.strftime("%Y-%m-%d %H:%M:%S"),
        "totales": totales,
        "cortes_turno": cortes_turno,
    }

    pdf_bytes = generar_pdf(payload_reporte)
    excel_bytes = generar_excel(payload_reporte)

    timestamp = datetime.now(mx_tz).strftime("%Y%m%d_%H%M%S")   # <-- CORREGIDO
    pdf_key = f"reportes/reporte_final_{corte_final_id}_{timestamp}.pdf"
    excel_key = f"reportes/reporte_final_{corte_final_id}_{timestamp}.xlsx"

    try:
        pdf_url = subir_a_s3(pdf_key, pdf_bytes, "application/pdf")
        excel_url = subir_a_s3(
            excel_key,
            excel_bytes,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        print("Error subiendo archivos a S3:", e)
        return jsonify({"error": "Error subiendo archivos a S3", "details": str(e)}), 500

    try:
        reporte_id = guardar_reporte_bd(corte_final_id, pdf_url, excel_url)
    except Exception as e:
        print("Error guardando reporte en BD:", e)
        return jsonify({"error": "Error guardando reporte en BD", "details": str(e)}), 500

    try:
        if NOTIFICACIONES_URL:
            resp_notify = requests.post(
                NOTIFICACIONES_URL,
                json={
                    "corte_final_id": corte_final_id,
                    "pdf_url": pdf_url,
                    "excel_url": excel_url,
                },
                timeout=5
            )
            print("Llamada a NOTIFICACIONES_URL status:",
                  resp_notify.status_code,
                  resp_notify.text)
        else:
            print("NOTIFICACIONES_URL no configurada, no se envía correo.")
    except Exception as e:
        print("Error llamando a NOTIFICACIONES_URL:", e)

    return jsonify({
        "message": "Reporte generado correctamente",
        "reporte_id": reporte_id,
        "corte_final_id": corte_final_id,
        "pdf_url": pdf_url,
        "excel_url": excel_url,
        "totales": totales,
        "num_cortes_turno": len(cortes_turno),
    }), 201


@app.route("/reportes", methods=["GET"])
def listar_reportes():
    sql = """
        SELECT r.id,
               r.corte_id,
               r.archivo_pdf_url,
               r.archivo_excel_url,
               r.fecha_generado,
               c.fecha_inicio AS fecha_corte_final
        FROM reportes r
        JOIN cortes c ON r.corte_id = c.id
        ORDER BY r.fecha_generado DESC
    """
    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute(sql)
            rows = cursor.fetchall()
    finally:
        conn.close()

    for r in rows:
        if isinstance(r.get("fecha_generado"), datetime):
            r["fecha_generado"] = r["fecha_generado"].strftime("%Y-%m-%d %H:%M:%S")
        if isinstance(r.get("fecha_corte_final"), datetime):
            r["fecha_corte_final"] = r["fecha_corte_final"].strftime("%Y-%m-%d %H:%M:%S")

    return jsonify(rows), 200


@app.route("/reportes/<int:reporte_id>", methods=["GET"])
def obtener_reporte(reporte_id):
    sql = """
        SELECT r.id,
               r.corte_id,
               r.archivo_pdf_url,
               r.archivo_excel_url,
               r.fecha_generado,
               c.fecha_inicio AS fecha_corte_final
        FROM reportes r
        JOIN cortes c ON r.corte_id = c.id
        WHERE r.id = %s
    """
    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute(sql, (reporte_id,))
            row = cursor.fetchone()
    finally:
        conn.close()

    if not row:
        return jsonify({"error": "Reporte no encontrado"}), 404

    if isinstance(row.get("fecha_generado"), datetime):
        row["fecha_generado"] = row["fecha_generado"].strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(row.get("fecha_corte_final"), datetime):
        row["fecha_corte_final"] = row["fecha_corte_final"].strftime("%Y-%m-%d %H:%M:%S")

    return jsonify(row), 200


# ==========================
# Local & Lambda handler
# ==========================

if __name__ == "__main__":
    port = int(os.getenv("PORT", "8004"))
    app.run(host="0.0.0.0", port=port, debug=True)

try:
    import awsgi

    def handler(event, context):
        return awsgi.response(app, event, context)
except ImportError:
    pass
